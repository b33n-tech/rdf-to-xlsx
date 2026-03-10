"""
Persée RDF → XLSX Converter
============================
Convertit les fichiers .rdf téléchargés depuis data.persee.fr en fichiers Excel.

Dépendances :
    pip install streamlit openpyxl lxml

Lancement :
    streamlit run persee_rdf_to_xlsx.py
"""

import io
import re
import xml.etree.ElementTree as ET
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st

# ──────────────────────────────────────────────────────────────────────────────
# Namespaces RDF utilisés par Persée
# ──────────────────────────────────────────────────────────────────────────────
NS = {
    "rdf":     "http://www.w3.org/1999/02/22-rdf-syntax-ns#",
    "rdfs":    "http://www.w3.org/2000/01/rdf-schema#",
    "dcterms": "http://purl.org/dc/terms/",
    "dc":      "http://purl.org/dc/elements/1.1/",
    "bibo":    "http://purl.org/ontology/bibo/",
    "biro":    "http://purl.org/spar/biro/",
    "foaf":    "http://xmlns.com/foaf/0.1/",
    "rdam":    "http://rdaregistry.info/Elements/m/",
    "rdau":    "http://rdaregistry.info/Elements/u/",
    "marcrel": "http://id.loc.gov/vocabulary/relators/",
    "cito":    "http://purl.org/spar/cito/",
    "skos":    "http://www.w3.org/2004/02/skos/core#",
    "owl":     "http://www.w3.org/2002/07/owl#",
    "schema":  "http://schema.org/",
    "bio":     "http://purl.org/vocab/bio/0.1/",
    "persee":  "http://data.persee.fr/ontology/persee-ontology/",
}

XML_LANG = "{http://www.w3.org/XML/1998/namespace}lang"
RDF_ABOUT = "{http://www.w3.org/1999/02/22-rdf-syntax-ns#}about"
RDF_RESOURCE = "{http://www.w3.org/1999/02/22-rdf-syntax-ns#}resource"


def _txt(el, path, default=""):
    """Retourne le texte d'un sous-élément ou default."""
    node = el.find(path, NS)
    if node is not None and node.text:
        return node.text.strip()
    return default


def _attr(el, path, attr=RDF_RESOURCE, default=""):
    node = el.find(path, NS)
    if node is not None:
        return node.get(attr, default)
    return default


def _all_txt(el, path):
    """Retourne une liste de textes pour des éléments multiples."""
    return [n.text.strip() for n in el.findall(path, NS) if n.text]


def _all_attr(el, path, attr=RDF_RESOURCE):
    return [n.get(attr, "") for n in el.findall(path, NS) if n.get(attr)]


# ──────────────────────────────────────────────────────────────────────────────
# Parsers par type
# ──────────────────────────────────────────────────────────────────────────────

def parse_collection(root):
    rows = []
    # On récupère tous les nœuds pertinents
    for tag in ("bibo:Collection", "bibo:Journal"):
        for el in root.findall(tag, NS):
            about = el.get(RDF_ABOUT, "")
            media = _txt(el, "rdam:mediaType")
            row = {
                "URI":               about,
                "Type_nœud":         tag.split(":")[1],
                "Média":             media,
                "Titre":             _txt(el, "dcterms:title") or _txt(el, "rdfs:label"),
                "Éditeur":           _txt(el, "dcterms:publisher"),
                "Identifiant_URL":   _txt(el, "dcterms:identifier"),
                "ISSN":              _attr(el, "bibo:issn"),
                "eISSN":             _attr(el, "bibo:eissn"),
                "Date_publication":  _txt(el, "rdam:dateOfPublication"),
                "Date_modification": _txt(el, "dcterms:modified"),
                "Couverture":        _txt(el, "dcterms:coverage"),
                "Licence":           _txt(el, "dcterms:license"),
                "Description_fr":    "",
                "Description_en":    "",
            }
            for desc in el.findall("dcterms:description", NS):
                lang = desc.get(XML_LANG, "")
                txt = (desc.text or "").strip()
                if lang == "fr":
                    row["Description_fr"] = txt
                elif lang == "en":
                    row["Description_en"] = txt
            rows.append(row)
    return rows


def parse_doc(root):
    rows = []
    for el in root.findall("bibo:Document", NS):
        about = el.get(RDF_ABOUT, "")
        if "#Web" not in about:
            continue  # on ne garde que la version Web (la plus complète)

        # Résumés par langue
        abstract_fr = abstract_en = ""
        for ab in el.findall("dcterms:abstract", NS):
            lang = ab.get(XML_LANG, "")
            txt = (ab.text or "").strip()
            if lang == "fr":
                abstract_fr = txt
            elif lang == "en":
                abstract_en = txt

        # Auteurs (URIs → extraire l'ID)
        creators = _all_attr(el, "dcterms:creator")
        auteurs = " | ".join(
            re.sub(r".*/authority/(\d+).*", r"\1", c) for c in creators
            if "authority" in c
        )

        # Citations
        cites      = " | ".join(_all_attr(el, "cito:cites"))
        cited_by   = " | ".join(_all_attr(el, "cito:isCitedBy"))

        row = {
            "URI":                   about,
            "Titre":                 _txt(el, "dcterms:title"),
            "Langue":                _txt(el, "dcterms:language"),
            "DOI":                   _txt(el, "bibo:doi"),
            "URL_Persée":            _txt(el, "dcterms:identifier"),
            "Date_publication_web":  _txt(el, "rdam:dateOfPublication"),
            "Date_publication_print":_txt(el, "persee:dateOfPrintPublication"),
            "Date_modification":     _txt(el, "dcterms:modified"),
            "Page_début":            _txt(el, "bibo:pageStart"),
            "Page_fin":              _txt(el, "bibo:pageEnd"),
            "Nb_pages":              _txt(el, "bibo:numPages"),
            "Numéro_Persée":         _attr(el, "dcterms:isPartOf"),
            "Auteurs_IDs":           auteurs,
            "Éditeur":               _txt(el, "dcterms:publisher"),
            "Licence":               _txt(el, "dcterms:license"),
            "Citation_bibliographique": _txt(el, "dcterms:bibliographicCitation"),
            "Résumé_fr":             abstract_fr,
            "Résumé_en":             abstract_en,
            "Cites":                 cites,
            "Cité_par":              cited_by,
        }
        rows.append(row)
    return rows


def parse_issue(root):
    rows = []
    for el in root.findall("bibo:Issue", NS):
        about = el.get(RDF_ABOUT, "")
        if "#Web" not in about:
            continue

        row = {
            "URI":                    about,
            "Titre":                  _txt(el, "dcterms:title"),
            "Label":                  _txt(el, "rdfs:label"),
            "Volume":                 _txt(el, "bibo:volume"),
            "Numéro":                 _txt(el, "bibo:issue"),
            "Langue":                 _txt(el, "dcterms:language"),
            "Date_publication_web":   _txt(el, "rdam:dateOfPublication"),
            "Date_publication_print": _txt(el, "persee:dateOfPrintPublication"),
            "Date_modification":      _txt(el, "dcterms:modified"),
            "URL_Persée":             _txt(el, "dcterms:identifier"),
            "Éditeur":                _txt(el, "dcterms:publisher"),
            "Licence":                _txt(el, "dcterms:license"),
            "Citation_bibliographique": _txt(el, "dcterms:bibliographicCitation"),
            "Thème":                  _txt(el, "foaf:focus"),
            "Image_couverture":       _attr(el, "foaf:depiction"),
            "Collection":             _attr(el, "dcterms:isPartOf"),
        }
        rows.append(row)
    return rows


def parse_persons(root):
    rows = []
    for el in root.findall("foaf:Person", NS):
        about = el.get(RDF_ABOUT, "")

        # Identifiants externes
        same_as = _all_attr(el, "owl:sameAs")
        idref = viaf = bnf = isni = dbpedia = ""
        for uri in same_as:
            if "idref.fr" in uri:
                idref = uri
            elif "viaf.org" in uri:
                viaf = uri
            elif "data.bnf.fr" in uri:
                bnf = uri
            elif "isni.org" in uri:
                isni = uri
            elif "dbpedia.org" in uri:
                dbpedia = uri

        pages = " | ".join(_all_attr(el, "foaf:page"))
        depictions = " | ".join(_all_attr(el, "foaf:depiction"))

        row = {
            "URI":          about,
            "Nom_complet":  _txt(el, "foaf:name"),
            "Prénom":       _txt(el, "foaf:givenName"),
            "Nom_famille":  _txt(el, "foaf:familyName"),
            "Date_naissance": _txt(el, "schema:birthDate"),
            "Date_décès":   _txt(el, "schema:deathDate"),
            "Biographie":   _txt(el, "bio:biography"),
            "Page_Persée":  "",
            "Pages_externes": pages,
            "IdRef":        idref,
            "VIAF":         viaf,
            "BnF":          bnf,
            "ISNI":         isni,
            "DBpedia":      dbpedia,
            "Images":       depictions,
        }
        # Isoler la page Persée
        for p in _all_attr(el, "foaf:page"):
            if "persee.fr/authority" in p:
                row["Page_Persée"] = p
        rows.append(row)
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# Générateur XLSX avec mise en forme
# ──────────────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL      = PatternFill("solid", fgColor="D9E1F2")
NORMAL_FILL   = PatternFill("solid", fgColor="FFFFFF")
BORDER_SIDE   = Side(style="thin", color="BDD7EE")
CELL_BORDER   = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE, bottom=BORDER_SIDE
)

# Largeurs de colonnes adaptées au contenu
COL_WIDTHS = {
    "URI": 55, "Titre": 45, "Résumé_fr": 60, "Résumé_en": 60,
    "Citation_bibliographique": 60, "Biographie": 60,
    "Description_fr": 60, "Description_en": 60,
    "Label": 50, "Thème": 40,
}
DEFAULT_WIDTH = 25


def build_xlsx(rows: list[dict], sheet_name: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    if not rows:
        ws["A1"] = "Aucune donnée trouvée."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(rows[0].keys())

    # En-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    ws.row_dimensions[1].height = 30

    # Données
    for row_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else NORMAL_FILL
        for col_idx, header in enumerate(headers, 1):
            val = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER
            cell.font = Font(name="Calibri", size=10)

    # Largeurs de colonnes
    for col_idx, header in enumerate(headers, 1):
        width = COL_WIDTHS.get(header, DEFAULT_WIDTH)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width

    # Figer la 1re ligne
    ws.freeze_panes = "A2"

    # Filtre auto
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# Interface Streamlit
# ──────────────────────────────────────────────────────────────────────────────

TYPES = {
    "collection — Métadonnées de la revue":        ("collection", parse_collection),
    "doc — Articles et documents":                 ("doc",        parse_doc),
    "issue — Numéros de la revue":                 ("issue",      parse_issue),
    "persons — Auteurs et personnes":              ("persons",    parse_persons),
}

st.set_page_config(
    page_title="Persée RDF → XLSX",
    page_icon="📚",
    layout="centered",
)

st.title("📚 Persée · RDF → XLSX")
st.markdown(
    "Convertissez les dumps RDF téléchargés depuis "
    "[data.persee.fr](https://data.persee.fr/explorer/demander-un-dump/) "
    "en fichier Excel exploitable."
)

st.divider()

col1, col2 = st.columns([2, 1])

with col1:
    rdf_type_label = st.selectbox(
        "**1. Sélectionnez le type de fichier**",
        options=list(TYPES.keys()),
        help="Chaque type de dump Persée a une structure différente. "
             "Sélectionnez celui qui correspond à votre fichier.",
    )

with col2:
    st.markdown("<br>", unsafe_allow_html=True)
    short_name = TYPES[rdf_type_label][0]
    st.info(f"Type détecté : **`{short_name}`**")

uploaded = st.file_uploader(
    "**2. Déposez votre fichier `.rdf`**",
    type=["rdf", "xml"],
    help="Fichier téléchargé depuis le portail Persée.",
)

if uploaded:
    st.success(f"Fichier chargé : `{uploaded.name}` ({uploaded.size:,} octets)")

    if st.button("🔄 Convertir en XLSX", type="primary", use_container_width=True):
        with st.spinner("Parsing du fichier RDF en cours…"):
            try:
                content = uploaded.read()
                # Supprimer la DTD (résolution d'entité externe bloque ET)
                content_str = content.decode("utf-8", errors="replace")
                content_str = re.sub(r"<!DOCTYPE\s+\w+\s*\[.*?\]>", "", content_str, flags=re.DOTALL)
                # Remplacer les entités XML non standards
                content_str = content_str.replace("&xsd;", "http://www.w3.org/2001/XMLSchema#")

                root = ET.fromstring(content_str)

                _, parser_fn = TYPES[rdf_type_label]
                rows = parser_fn(root)

                if not rows:
                    st.warning(
                        "⚠️ Aucune ligne extraite. Vérifiez que le type sélectionné "
                        "correspond bien au fichier uploadé."
                    )
                else:
                    xlsx_bytes = build_xlsx(rows, sheet_name=short_name)

                    # Nom de fichier de sortie
                    base = uploaded.name.replace(".rdf", "").replace(".xml", "")
                    out_name = f"{base}.xlsx"

                    st.success(f"✅ **{len(rows):,} lignes** extraites avec succès !")

                    # Aperçu
                    with st.expander("👁️ Aperçu des 5 premières lignes"):
                        import pandas as pd
                        df = pd.DataFrame(rows[:5])
                        st.dataframe(df, use_container_width=True)

                    st.download_button(
                        label=f"⬇️ Télécharger `{out_name}`",
                        data=xlsx_bytes,
                        file_name=out_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )

            except ET.ParseError as e:
                st.error(f"❌ Erreur de parsing XML : {e}")
            except Exception as e:
                st.error(f"❌ Erreur inattendue : {e}")
                raise e

st.divider()
st.caption(
    "**Types de fichiers Persée pris en charge :**  "
    "`collection` · `doc` · `issue` · `persons`  \n"
    "Chaque revue téléchargée produit ces 4 fichiers avec le même nommage "
    "`PERSEE_{revue}_{type}_{date}.rdf`."
)
